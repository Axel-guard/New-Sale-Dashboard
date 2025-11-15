#!/usr/bin/env python3
"""
Excel Data Import Script for AxelGuard Dashboard
Imports data from 3 sheets: Dispatch, QC Status, and Inventory
"""

import openpyxl
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

# Database path (local D1 database)
DB_PATH = "/home/user/webapp/.wrangler/state/v3/d1/miniflare-D1DatabaseObject/a4cbf95b06cc05ac18912e42ea1dd3c229ea877895f964b2fcd2b1a46ff17dbc.sqlite"
EXCEL_FILE = "/home/user/uploaded_files/Inventory QC.xlsx"

def format_date(value):
    """Convert Excel date to SQLite date format"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, str):
        try:
            dt = datetime.strptime(value, '%Y-%m-%d')
            return dt.strftime('%Y-%m-%d')
        except:
            return None
    return None

def clean_value(value):
    """Clean and prepare value for database insertion"""
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip() if value.strip() else None
    return value

def import_inventory_sheet(wb, conn):
    """Import data from Inventory sheet"""
    print("\n" + "="*60)
    print("📦 IMPORTING INVENTORY DATA")
    print("="*60)
    
    ws = wb['Inventory']
    cursor = conn.cursor()
    
    # Clear existing data
    cursor.execute("DELETE FROM inventory")
    conn.commit()
    print("✅ Cleared existing inventory data")
    
    success_count = 0
    error_count = 0
    skip_count = 0
    
    for row_idx in range(2, ws.max_row + 1):
        row = ws[row_idx]
        
        # Extract data
        s_no = clean_value(row[0].value)
        in_date = format_date(row[1].value)
        model_name = clean_value(row[2].value)
        device_serial_no = clean_value(row[3].value)
        dispatch_date = format_date(row[4].value)
        cust_code = clean_value(row[5].value)
        sale_date = format_date(row[6].value)
        customer_name = clean_value(row[7].value)
        cust_city = clean_value(row[8].value)
        cust_mobile = clean_value(row[9].value)
        dispatch_reason = clean_value(row[10].value)
        warranty_provide = clean_value(row[11].value)
        old_serial_no = clean_value(row[12].value)
        license_renew_time = format_date(row[13].value)
        user_id = clean_value(row[14].value)
        password = clean_value(row[15].value)
        account_activation_date = format_date(row[16].value)
        account_expiry_date = format_date(row[17].value)
        
        # Skip if no serial number
        if not device_serial_no:
            skip_count += 1
            continue
        
        # Determine status based on dispatch_date
        if dispatch_date:
            status = 'Dispatched'
        else:
            status = 'In Stock'
        
        try:
            cursor.execute('''
                INSERT INTO inventory (
                    serial_number, in_date, model_name, device_serial_no,
                    dispatch_date, cust_code, sale_date, customer_name,
                    cust_city, cust_mobile, dispatch_reason, warranty_provide,
                    old_serial_no, license_renew_time, user_id, password,
                    account_activation_date, account_expiry_date, status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                s_no, in_date, model_name, device_serial_no,
                dispatch_date, cust_code, sale_date, customer_name,
                cust_city, cust_mobile, dispatch_reason, warranty_provide,
                old_serial_no, license_renew_time, user_id, password,
                account_activation_date, account_expiry_date, status
            ))
            success_count += 1
            
            if success_count % 500 == 0:
                print(f"  ⏳ Processed {success_count} inventory records...")
                conn.commit()
                
        except sqlite3.IntegrityError as e:
            error_count += 1
            if error_count <= 5:
                print(f"  ⚠️  Row {row_idx}: Duplicate serial number {device_serial_no}")
        except Exception as e:
            error_count += 1
            if error_count <= 5:
                print(f"  ❌ Row {row_idx}: Error - {str(e)}")
    
    conn.commit()
    print(f"\n✅ Inventory Import Complete:")
    print(f"   • Success: {success_count} records")
    print(f"   • Errors: {error_count} records")
    print(f"   • Skipped: {skip_count} records (no serial number)")
    
    return success_count

def import_dispatch_sheet(wb, conn):
    """Import data from Dispatch sheet"""
    print("\n" + "="*60)
    print("🚚 IMPORTING DISPATCH DATA")
    print("="*60)
    
    ws = wb['Dispatch']
    cursor = conn.cursor()
    
    # Clear existing dispatch records
    cursor.execute("DELETE FROM dispatch_records")
    conn.commit()
    print("✅ Cleared existing dispatch records")
    
    success_count = 0
    error_count = 0
    skip_count = 0
    
    for row_idx in range(2, ws.max_row + 1):
        row = ws[row_idx]
        
        # Extract data
        s_no = clean_value(row[0].value)
        device_serial_no = clean_value(row[1].value)
        device_name = clean_value(row[2].value)
        qc_status = clean_value(row[3].value) or 'Pending'
        dispatch_reason = clean_value(row[4].value)
        order_id = clean_value(row[5].value)
        cust_code = clean_value(row[6].value)
        customer_name = clean_value(row[7].value)
        company_name = clean_value(row[8].value)
        dispatch_date = format_date(row[9].value)
        courier_company = clean_value(row[10].value)
        dispatch_method = clean_value(row[11].value)
        tracking_id = clean_value(row[12].value)
        
        # Skip if no serial number or dispatch date
        if not device_serial_no or not dispatch_date:
            skip_count += 1
            continue
        
        # Find inventory_id for this serial number
        cursor.execute(
            "SELECT id FROM inventory WHERE device_serial_no = ?",
            (str(device_serial_no),)
        )
        result = cursor.fetchone()
        
        if not result:
            # Create inventory record if doesn't exist
            try:
                cursor.execute('''
                    INSERT INTO inventory (
                        device_serial_no, model_name, status,
                        dispatch_date, customer_name, cust_code
                    ) VALUES (?, ?, 'Dispatched', ?, ?, ?)
                ''', (str(device_serial_no), device_name or 'Unknown', dispatch_date, customer_name, cust_code))
                inventory_id = cursor.lastrowid
            except:
                skip_count += 1
                continue
        else:
            inventory_id = result[0]
        
        try:
            cursor.execute('''
                INSERT INTO dispatch_records (
                    serial_number, inventory_id, device_serial_no, dispatch_date,
                    customer_name, customer_code, dispatch_reason, courier_name,
                    tracking_number, dispatched_by, order_id, qc_status,
                    dispatch_method, company_name
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                s_no, inventory_id, str(device_serial_no), dispatch_date,
                customer_name or 'Unknown', cust_code, dispatch_reason,
                courier_company, tracking_id, 'System Import', order_id,
                qc_status, dispatch_method, company_name
            ))
            success_count += 1
            
            if success_count % 500 == 0:
                print(f"  ⏳ Processed {success_count} dispatch records...")
                conn.commit()
                
        except Exception as e:
            error_count += 1
            if error_count <= 5:
                print(f"  ❌ Row {row_idx}: Error - {str(e)}")
    
    conn.commit()
    print(f"\n✅ Dispatch Import Complete:")
    print(f"   • Success: {success_count} records")
    print(f"   • Errors: {error_count} records")
    print(f"   • Skipped: {skip_count} records")
    
    return success_count

def import_qc_sheet(wb, conn):
    """Import data from QC Status sheet"""
    print("\n" + "="*60)
    print("✅ IMPORTING QC STATUS DATA")
    print("="*60)
    
    ws = wb['QC Status']
    cursor = conn.cursor()
    
    # Clear existing QC records
    cursor.execute("DELETE FROM quality_check")
    conn.commit()
    print("✅ Cleared existing QC records")
    
    success_count = 0
    error_count = 0
    skip_count = 0
    
    for row_idx in range(2, ws.max_row + 1):
        row = ws[row_idx]
        
        # Extract data
        s_no = clean_value(row[0].value)
        qc_date = format_date(row[1].value)
        serial_number = clean_value(row[2].value)
        device_type = clean_value(row[3].value)
        camera_quality = clean_value(row[4].value)
        sd_connectivity = clean_value(row[5].value)
        all_ch_status = clean_value(row[6].value)
        network_connectivity = clean_value(row[7].value)
        gps_qc = clean_value(row[8].value)
        sim_slot_qc = clean_value(row[9].value)
        online_qc = clean_value(row[10].value)
        monitor_qc = clean_value(row[11].value)
        final_qc_status = clean_value(row[12].value)
        ip_address_update = clean_value(row[13].value)
        final_remarks = clean_value(row[14].value)
        
        # Skip if no serial number
        if not serial_number:
            skip_count += 1
            continue
        
        # Find inventory_id
        cursor.execute(
            "SELECT id FROM inventory WHERE device_serial_no = ?",
            (str(serial_number),)
        )
        result = cursor.fetchone()
        
        if not result:
            # Create inventory record if doesn't exist
            try:
                cursor.execute('''
                    INSERT INTO inventory (
                        device_serial_no, model_name, status
                    ) VALUES (?, ?, 'Quality Check')
                ''', (str(serial_number), device_type or 'Unknown'))
                inventory_id = cursor.lastrowid
            except:
                skip_count += 1
                continue
        else:
            inventory_id = result[0]
        
        # Determine final status
        if final_qc_status and 'pass' in final_qc_status.lower():
            qc_status = 'Pass'
        elif final_qc_status and 'fail' in final_qc_status.lower():
            qc_status = 'Fail'
        else:
            qc_status = 'Pending'
        
        # Build detailed test results
        test_results_parts = []
        if camera_quality:
            test_results_parts.append(f"Camera: {camera_quality}")
        if sd_connectivity:
            test_results_parts.append(f"SD Card: {sd_connectivity}")
        if all_ch_status:
            test_results_parts.append(f"All Channels: {all_ch_status}")
        if network_connectivity:
            test_results_parts.append(f"Network: {network_connectivity}")
        if gps_qc:
            test_results_parts.append(f"GPS: {gps_qc}")
        if sim_slot_qc:
            test_results_parts.append(f"SIM Slot: {sim_slot_qc}")
        if online_qc:
            test_results_parts.append(f"Online: {online_qc}")
        if monitor_qc:
            test_results_parts.append(f"Monitor: {monitor_qc}")
        if ip_address_update:
            test_results_parts.append(f"IP Address: {ip_address_update}")
        
        test_results = " | ".join(test_results_parts) if test_results_parts else "No test details"
        
        try:
            cursor.execute('''
                INSERT INTO quality_check (
                    serial_number, inventory_id, device_serial_no, check_date,
                    checked_by, test_results, pass_fail, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                s_no, inventory_id, str(serial_number),
                qc_date or datetime.now().strftime('%Y-%m-%d'),
                'System Import', test_results, qc_status, final_remarks
            ))
            success_count += 1
            
            if success_count % 500 == 0:
                print(f"  ⏳ Processed {success_count} QC records...")
                conn.commit()
                
        except Exception as e:
            error_count += 1
            if error_count <= 5:
                print(f"  ❌ Row {row_idx}: Error - {str(e)}")
    
    conn.commit()
    print(f"\n✅ QC Import Complete:")
    print(f"   • Success: {success_count} records")
    print(f"   • Errors: {error_count} records")
    print(f"   • Skipped: {skip_count} records")
    
    return success_count

def main():
    print("\n" + "="*60)
    print("🚀 STARTING EXCEL DATA IMPORT")
    print("="*60)
    
    # Check if files exist
    if not Path(EXCEL_FILE).exists():
        print(f"❌ Error: Excel file not found at {EXCEL_FILE}")
        sys.exit(1)
    
    if not Path(DB_PATH).exists():
        print(f"❌ Error: Database file not found at {DB_PATH}")
        print("   Run: npm run dev first to create the database")
        sys.exit(1)
    
    # Load Excel file
    print(f"\n📂 Loading Excel file...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    print(f"✅ Loaded {len(wb.sheetnames)} sheets: {wb.sheetnames}")
    
    # Connect to database
    print(f"\n🔌 Connecting to database...")
    conn = sqlite3.connect(DB_PATH)
    print(f"✅ Connected to local D1 database")
    
    try:
        # Import data in sequence
        total_inventory = import_inventory_sheet(wb, conn)
        total_dispatch = import_dispatch_sheet(wb, conn)
        total_qc = import_qc_sheet(wb, conn)
        
        # Summary
        print("\n" + "="*60)
        print("🎉 IMPORT COMPLETED SUCCESSFULLY")
        print("="*60)
        print(f"\n📊 Summary:")
        print(f"   • Inventory Records: {total_inventory}")
        print(f"   • Dispatch Records: {total_dispatch}")
        print(f"   • QC Records: {total_qc}")
        print(f"   • Total Records: {total_inventory + total_dispatch + total_qc}")
        
        print("\n✅ All data has been imported successfully!")
        print("\n🌐 Next steps:")
        print("   1. Restart the development server: pm2 restart webapp")
        print("   2. Visit: https://3000-id7zgaopnm7accybu066c-b32ec7bb.sandbox.novita.ai")
        print("   3. Navigate to Inventory Management to see your data")
        
    except Exception as e:
        print(f"\n❌ Fatal Error: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        conn.close()
        wb.close()
        print("\n🔒 Database connection closed")

if __name__ == "__main__":
    main()
